'''
Os漏洞数据:
1.linux特殊
2.其他操作系统

目标网站:
https://syzkaller.appspot.com/upstream/subsystems

'''
import os
import re
import requests
from lxml import etree
from urllib.parse import urljoin
from loguru import logger
from openpyxl import workbook

'''
表一:
Title
Tag
Date
Sample
Crashes
'''

'''
表二:
Name
Open
Fixed

'''


class SyzbotSpider():
    class Method:
        GET = 'GET'
        POST = 'POST'

    def __init__(self):
        self.max_row = None
        self.diname = 'syzbot_data'  # 保存文件的目录
        self.re_obj = re.compile('Status: (.*?)<br>', re.S)
        self.item_list = [
            {'url': 'https://syzkaller.appspot.com/upstream/fixed',  # Subsystems 特征: upstream
             'file_name': 'linux.xlsx'
             },

            # {'url': 'https://syzkaller.appspot.com/linux-5.15/fixed',
            #  'file_name': 'linux_5_15.xlsx'
            #  },

            # {'url': 'https://syzkaller.appspot.com/linux-6.1/fixed',
            #  'file_name': 'linux_6_1.xlsx'
            #  },

            # {'url': 'https://syzkaller.appspot.com/netbsd/fixed',
            #  'file_name': 'netbsd.xlsx'
            #  },
            # {'url': 'https://syzkaller.appspot.com/openbsd/fixed',
            #  'file_name': 'openbsd.xlsx'
            #  },
            # {'url': 'https://syzkaller.appspot.com/gvisor/fixed',
            #  'file_name': 'gvisor.xlsx'
            #  },

        ]
        self.headers = {
            "authority": "syzkaller.appspot.com",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "zh-CN,zh;q=0.9",
            "cache-control": "no-cache",
            "pragma": "no-cache",
            "referer": "https://syzkaller.appspot.com/android-5-10/fixed",
            "sec-ch-ua": "\"Not A(Brand\";v=\"99\", \"Google Chrome\";v=\"121\", \"Chromium\";v=\"121\"",
            "sec-ch-ua-arch": "\"x86\"",
            "sec-ch-ua-bitness": "\"64\"",
            "sec-ch-ua-full-version-list": "\"Not A(Brand\";v=\"99.0.0.0\", \"Google Chrome\";v=\"121.0.6167.185\", \"Chromium\";v=\"121.0.6167.185\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-model": "\"\"",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-ch-ua-platform-version": "\"10.0.0\"",
            "sec-ch-ua-wow64": "?0",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        }
        self.cookies = {
            "syzkaller": "eyJuYW1lc3BhY2UiOiJhbmRyb2lkLTUtMTAifQ==",
            "_ga": "GA1.3.1559605917.1708171024",
            "_gid": "GA1.3.1794547718.1708171024",
            "_gat_gtag_UA_116117799_1": "1"
        }
        self.base_dir = os.path.dirname(os.path.abspath(__file__))

        if not os.path.exists(self.diname):
            os.makedirs(self.diname)

    def get_date_info(self, data_str):
        """ 日期数据提取 """
        tree = etree.HTML(data_str)
        return ''.join(tree.xpath('//a/text()'))

    def _request(self, method: str = 'GET', *args, **kwargs):
        """请求封装"""
        error_count = 0
        # 是否挂代理
        # if not self.proxies:
        #     headers, self.proxies = self.get_proxies(self.headers)
        kwargs.update(
            {
                'headers': self.headers,
                'cookies': self.cookies,
                'proxies': {
                    'http': 'http://127.0.0.1:7890',
                    'https': 'http://127.0.0.1:7890',
                }
            }
        )
        while True:
            try:
                if error_count > 10:
                    return

                if method == SyzbotSpider.Method.GET:
                    resp = requests.get(**kwargs)
                else:
                    resp = requests.post(**kwargs)

                return resp

            except Exception as e:
                logger.error('_request error {}'.format(e))
                error_count += 1

    def write_excel_bar(self, sheet):
        """
        excel表头写入
        :return:
        """
        sheet.cell(1, 1).value = "Title"
        sheet.cell(1, 2).value = "Tag"
        sheet.cell(1, 3).value = "Date"
        sheet.cell(1, 4).value = "Sample"
        sheet.cell(1, 5).value = "Crashes"
        sheet.cell(1, 6).value = "Href"

    def write_excel_bar2(self, sheet):
        """
        excel表头写入
        :return:
        """
        sheet.cell(1, 1).value = "Name"
        sheet.cell(1, 2).value = "Open"
        sheet.cell(1, 3).value = "Fixed"
        sheet.cell(1, 4).value = "Href"

    def write_excel_data(self, data_dict, sheet):
        """ 写入核心数据 """
        logger.info(data_dict)
        sheet.cell(self.max_row, 1).value = data_dict['title']
        sheet.cell(self.max_row, 2).value = data_dict['tag']
        sheet.cell(self.max_row, 3).value = data_dict['date']
        sheet.cell(self.max_row, 4).value = data_dict['sample']
        sheet.cell(self.max_row, 5).value = data_dict['crashes']
        sheet.cell(self.max_row, 6).value = data_dict['href']

    def write_excel_data2(self, data_dict, sheet):
        """ 写入核心数据 """
        logger.info(data_dict)
        sheet.cell(self.max_row, 1).value = data_dict['name']
        sheet.cell(self.max_row, 2).value = data_dict['open']
        sheet.cell(self.max_row, 3).value = data_dict['fixed']
        sheet.cell(self.max_row, 4).value = data_dict['href']

    def get_detail(self, href, data_dict, sheet):
        """ 详情页采集 """

        source_data = self._request(**{
            'url': href,
        }).text
        result = self.re_obj.search(source_data)
        if not result:
            logger.error(f'{href} 详情页数据提取错误!')
            # raise Exception('详情页数据提取错误!')
            data_dict['date'] = result.group(1)
        else:
            data_dict['date'] = self.get_date_info(result.group(1))

        # sample
        tree = etree.HTML(source_data)
        sample = ''.join(tree.xpath('//div[@id="crash_div"]/pre//text()'))

        data_dict['sample'] = sample

        crashes = ''.join(tree.xpath('//table[@class="list_table"]/caption/text()'))

        data_dict['crashes'] = re.findall('\d+', crashes, )[0]
        data_dict['href'] = href

        # 数据存储
        self.write_excel_data(data_dict, sheet)

    def runner(self, ):
        """ 程序入口"""

        href = None
        for item in self.item_list:
            self.max_row = 2
            file_name = item['file_name']
            target_url = item['url']
            logger.info(f'正在为您采集{file_name} & {target_url}'.center(50, '-'))

            source_data = self._request(**{
                'url': target_url,
            }).text
            wb = workbook.Workbook()
            sheet = wb["Sheet"]
            self.write_excel_bar(sheet)

            tree = etree.HTML(source_data)
            # print(source_data)
            trs = tree.xpath("//table[@class='list_table']/tbody/tr")
            logger.success(f'检测到该链接存在数据: 【{len(trs)}】条')
            file_path = os.path.join(self.base_dir, self.diname, file_name)
            for idx, tr in enumerate(trs):
                try:
                    title = ''.join(tr.xpath("./td[@class='title']/a/text()"))
                    tag = '  '.join(tree.xpath(".//span[@class='bug-label']/a/text()"))
                    href = urljoin(target_url, ''.join(tr.xpath("./td[@class='title']/a/@href")))
                    data_dict = {
                        'title': title,
                        'tag': tag,
                    }
                    # 采集详情页数据 sample crashes status
                    self.get_detail(href, data_dict, sheet)
                    self.max_row += 1
                    if idx % 20 == 1:
                        logger.info(f'正在保存: 【{file_path}】')
                        wb.save(file_path)
                    # break
                except Exception as e:
                    logger.info(f'error: {e}  {href}')
                    continue

            wb.save(file_path)


            # 针对linux条件 额外采集>>>
            # if 'upstream' in target_url:
            #     url2 = 'https://syzkaller.appspot.com/upstream/subsystems'
            #     file_name2 = f'subsystems_{item["file_name"]}'
            #     logger.info(f'正在为您采集{file_name2} & {url2}'.center(50, '-'))
            #     self.max_row = 2
            #     wb2 = workbook.Workbook()
            #     sheet2 = wb2["Sheet"]
            #     self.write_excel_bar2(sheet2)
            #     source_data2 = self._request(**{
            #         'url': url2,
            #     }).text
            #     tree = etree.HTML(source_data2)
            #     trs = tree.xpath('//table[@class="list_table"]/tbody/tr')
            #     logger.success(f'检测到该链接存在数据: 【{len(trs)}】条')
            #     for tr in trs:
            #         name = ''.join(tr.xpath('./td[1]//text()'))
            #         href = urljoin(url2, ''.join(tr.xpath('./td[1]/a/@ref')))
            #         open = ''.join(tr.xpath('./td[3]//text()'))
            #         fixed = ''.join(tr.xpath('./td[4]//text()'))
            #         data_dict2 = {
            #             'href': href,
            #             'name': name,
            #             'open': open,
            #             'fixed': fixed,
            #         }
            #         # 数据存储
            #         self.write_excel_data2(data_dict2, sheet2)
            #         self.max_row += 1
            #     file_path2 = os.path.join(self.base_dir, self.diname, file_name2)
            #     logger.info(f'正在保存: 【{file_path2}】')
            #     wb2.save(file_path2)
        else:
            logger.success('恭喜您、数据采集完成!')


if __name__ == '__main__':
    obj = SyzbotSpider()
    obj.runner()
    # obj.test()
